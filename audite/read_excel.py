import pandas as pd
import json
import os
from openpyxl import load_workbook

try:
    # Imprimir el directorio actual y listar archivos
    print("Directorio actual:", os.getcwd())
    print("\nArchivos en el directorio:")
    for file in os.listdir():
        if file.endswith('.xlsx'):
            print(f"- {file}")

    excel_file = 'BD-AGRO.xlsx'
    print(f"\nIntentando leer: {excel_file}")
    
    # Primero, obtener información sobre las hojas del Excel
    wb = load_workbook(filename=excel_file, read_only=True)
    print("\nHojas encontradas en el Excel:")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")
    
    # Intentar leer cada hoja
    for sheet_name in wb.sheetnames:
        print(f"\nLeyendo hoja: {sheet_name}")
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        print(f"Información de la hoja {sheet_name}:")
        print(f"Número de filas: {len(df)}")
        print(f"Número de columnas: {len(df.columns)}")
        print("\nColumnas encontradas:")
        for col in df.columns:
            print(f"- {col}")
        
        # Guardar información de esta hoja
        columns_info = {
            'nombres': list(df.columns),
            'tipos': {col: str(df[col].dtype) for col in df.columns}
        }

        sample_data = df.head().to_dict('records')

        output = {
            'hoja': sheet_name,
            'estructura': columns_info,
            'muestra_datos': sample_data
        }

        output_file = f'excel_analysis_{sheet_name}.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        print(f"\nAnálisis guardado en: {output_file}")

except Exception as e:
    print(f"\nError al procesar el archivo: {str(e)}") 